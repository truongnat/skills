#!/usr/bin/env python3
"""Convert JP-style / merged-cell Excel documents to HTML + Markdown."""

from __future__ import annotations

import argparse
import html
import json
import re
import sys
from pathlib import Path
from typing import Any


def _require_openpyxl() -> None:
    try:
        import openpyxl  # noqa: F401
    except ImportError as exc:
        raise SystemExit(
            "Missing openpyxl. Run this skill's scripts/setup.py, then invoke "
            "via the skill .venv python."
        ) from exc


def _safe_name(title: str) -> str:
    return re.sub(r"[^\w\-]+", "_", title.strip())[:60] or "sheet"


def _cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).replace("\r\n", "\n").replace("\r", "\n").strip()


def _merge_maps(ws: Any) -> tuple[dict[tuple[int, int], tuple[int, int]], set[tuple[int, int]]]:
    origin: dict[tuple[int, int], tuple[int, int]] = {}
    covered: set[tuple[int, int]] = set()
    for rng in ws.merged_cells.ranges:
        rs = rng.max_row - rng.min_row + 1
        cs = rng.max_col - rng.min_col + 1
        origin[(rng.min_row, rng.min_col)] = (rs, cs)
        for r in range(rng.min_row, rng.max_row + 1):
            for c in range(rng.min_col, rng.max_col + 1):
                if (r, c) != (rng.min_row, rng.min_col):
                    covered.add((r, c))
    return origin, covered


def _used_bounds(ws: Any, scan_rows: int, scan_cols: int) -> tuple[int, int]:
    from openpyxl.cell.cell import MergedCell

    max_r, max_c = 1, 1
    row_limit = min(ws.max_row or 1, scan_rows)
    col_limit = min(ws.max_column or 1, scan_cols)
    for r in range(1, row_limit + 1):
        for c in range(1, col_limit + 1):
            cell = ws.cell(r, c)
            if isinstance(cell, MergedCell):
                continue
            text = _cell_text(cell.value)
            if not text:
                continue
            max_r = max(max_r, r)
            max_c = max(max_c, c)
    for rng in ws.merged_cells.ranges:
        if rng.min_row <= max_r + 2 and rng.min_col <= max_c + 2:
            max_r = max(max_r, min(rng.max_row, scan_rows))
            max_c = max(max_c, min(rng.max_col, scan_cols))
    return max_r, max_c


def _sheet_strategy(title: str, merges: int, max_col: int) -> str:
    t = title.strip()
    if "レイアウト" in t or "layout" in t.lower():
        return "layout-asset"
    if merges >= 200 or max_col >= 80:
        return "table-dense"
    if "表紙" in t or "cover" in t.lower():
        return "cover-kv"
    if merges >= 20:
        return "merged-doc"
    return "simple-grid"


def classify_workbook(path: Path) -> dict[str, Any]:
    _require_openpyxl()
    from openpyxl import load_workbook

    wb = load_workbook(path, data_only=False)
    sheets: list[dict[str, Any]] = []
    for ws in wb.worksheets:
        merges = len(list(ws.merged_cells.ranges))
        used_r, used_c = _used_bounds(ws, scan_rows=120, scan_cols=80)
        sheets.append(
            {
                "title": ws.title,
                "declared": f"{ws.max_row}x{ws.max_column}",
                "merges": merges,
                "used_approx": f"{used_r}x{used_c}",
                "strategy": _sheet_strategy(ws.title, merges, used_c),
            }
        )
    wb.close()
    return {
        "format": "excel-doc-convert",
        "operation": "classify",
        "source": str(path),
        "sheet_count": len(sheets),
        "merge_total": sum(s["merges"] for s in sheets),
        "sheets": sheets,
    }


def sheet_to_html(ws: Any, max_r: int, max_c: int, title: str) -> str:
    origin, covered = _merge_maps(ws)
    parts = [
        "<!DOCTYPE html><html><head><meta charset='utf-8'>",
        f"<title>{html.escape(title)}</title>",
        "<style>table{border-collapse:collapse;font:12px/1.4 sans-serif}"
        "td{border:1px solid #888;padding:4px 6px;vertical-align:top;max-width:280px}"
        "caption{font-weight:600;margin:8px 0;text-align:left}</style></head><body>",
        f"<caption>{html.escape(title)} — export {max_r}x{max_c}, "
        f"merges={len(list(ws.merged_cells.ranges))}</caption>",
        "<table>",
    ]
    for r in range(1, max_r + 1):
        parts.append("<tr>")
        for c in range(1, max_c + 1):
            if (r, c) in covered:
                continue
            rs, cs = origin.get((r, c), (1, 1))
            rs = min(rs, max_r - r + 1)
            cs = min(cs, max_c - c + 1)
            val = _cell_text(ws.cell(r, c).value)
            attrs: list[str] = []
            if rs > 1:
                attrs.append(f'rowspan="{rs}"')
            if cs > 1:
                attrs.append(f'colspan="{cs}"')
            a = (" " + " ".join(attrs)) if attrs else ""
            parts.append(f"<td{a}>{html.escape(val).replace(chr(10), '<br>')}</td>")
        parts.append("</tr>")
    parts.append("</table></body></html>")
    return "\n".join(parts)


def _resolved_value(ws: Any, r: int, c: int) -> str:
    from openpyxl.cell.cell import MergedCell

    cell = ws.cell(r, c)
    if not isinstance(cell, MergedCell) and cell.value is not None:
        return _cell_text(cell.value)
    for rng in ws.merged_cells.ranges:
        if rng.min_row <= r <= rng.max_row and rng.min_col <= c <= rng.max_col:
            return _cell_text(ws.cell(rng.min_row, rng.min_col).value)
    return ""


def md_cover_kv(ws: Any, max_r: int, max_c: int, title: str) -> str:
    origin, covered = _merge_maps(ws)
    lines = [f"# {title.strip()}", "", "## Fields", "", "| 項目 | 内容 |", "| --- | --- |"]
    seen: set[tuple[str, str]] = set()
    # Stamp / signature row labels that tend to chain falsely — keep but mark later
    for r in range(1, max_r + 1):
        for c in range(1, max_c):
            if (r, c) in covered:
                continue
            left = _cell_text(ws.cell(r, c).value)
            if not left or len(left) > 40:
                continue
            right = ""
            start = c + 1
            if (r, c) in origin:
                start = c + origin[(r, c)][1]
            for cc in range(start, min(c + 12, max_c + 1)):
                if (r, cc) in covered:
                    continue
                t = _cell_text(ws.cell(r, cc).value)
                if t:
                    right = t
                    break
            if not right or left == right:
                continue
            key = (left, right)
            if key in seen:
                continue
            seen.add(key)
            lines.append(
                f"| {left.replace('|', '\\|')} | "
                f"{right.replace('|', '\\|').replace(chr(10), ' / ')} |"
            )
    lines.append("")
    lines.append(f"_pairs={len(seen)}_")
    return "\n".join(lines)


def md_blocks(ws: Any, max_r: int, max_c: int, title: str, limit_rows: int = 60) -> str:
    origin, covered = _merge_maps(ws)
    lines = [f"# {title.strip()}", "", "## Content", ""]
    count = 0
    for r in range(1, max_r + 1):
        texts: list[str] = []
        for c in range(1, max_c + 1):
            if (r, c) in covered:
                continue
            t = _cell_text(ws.cell(r, c).value)
            if not t:
                continue
            rs, cs = origin.get((r, c), (1, 1))
            span = f" [{rs}x{cs}]" if rs * cs > 1 else ""
            texts.append(t + span)
        if not texts:
            continue
        if len(texts) == 1:
            lines.append(texts[0])
            lines.append("")
        else:
            lines.append("- " + " | ".join(texts[:14]))
        count += 1
        if count >= limit_rows:
            lines.append("")
            lines.append("_… truncated_")
            break
    return "\n".join(lines)


def md_control_table(ws: Any, title: str, max_rows: int = 200) -> str | None:
    """Detect № / コントロール名 / コントロールID style tables (JP screen specs)."""
    header_r = None
    for r in range(1, min(60, (ws.max_row or 1) + 1)):
        vals = [_resolved_value(ws, r, c) for c in range(1, min(80, (ws.max_column or 1) + 1))]
        joined = " ".join(vals)
        if "コントロール名" in joined and "コントロールID" in joined:
            header_r = r
            break
    if header_r is None:
        return None

    labels: dict[str, int] = {}
    for c in range(1, min(100, (ws.max_column or 1) + 1)):
        v = _resolved_value(ws, header_r, c)
        if not v:
            continue
        if v.startswith("№") or v in {"No", "NO", "番号"}:
            labels.setdefault("no", c)
        elif "コントロール名" in v:
            labels.setdefault("name", c)
        elif "コントロールID" in v:
            labels.setdefault("id", c)
    if "name" not in labels or "id" not in labels:
        return None

    lines = [
        f"# {title.strip()}",
        "",
        "| № | コントロール名 | コントロールID |",
        "| --- | --- | --- |",
    ]
    rows = 0
    for r in range(header_r + 1, min(header_r + max_rows, (ws.max_row or 1) + 1)):
        no = _resolved_value(ws, r, labels.get("no", 0)) if labels.get("no") else ""
        name = _resolved_value(ws, r, labels["name"])
        cid = _resolved_value(ws, r, labels["id"])
        if not any([no, name, cid]):
            continue
        # Section banner: same text across cells
        if name and name == cid and (not no or no == name):
            lines.extend(["", f"### {name}", "", "| № | コントロール名 | コントロールID |", "| --- | --- | --- |"])
            continue
        if not name and not cid:
            continue
        lines.append(
            f"| {no.replace('|', '\\|')} | {name.replace('|', '\\|')} | {cid.replace('|', '\\|')} |"
        )
        rows += 1
    if rows == 0:
        return None
    lines.append("")
    lines.append(f"_rows={rows}_")
    return "\n".join(lines)


def convert_workbook(
    path: Path,
    out_dir: Path,
    *,
    sheet_filter: list[str] | None,
    max_rows: int,
    max_cols: int,
) -> dict[str, Any]:
    _require_openpyxl()
    from openpyxl import load_workbook

    out_dir.mkdir(parents=True, exist_ok=True)
    wb = load_workbook(path, data_only=False)
    wanted = {s.strip() for s in sheet_filter} if sheet_filter else None
    results: list[dict[str, Any]] = []
    limitations: list[str] = [
        "Shapes, charts, comments, and drawings are not converted.",
        "Declared sheet size may be template padding; export uses used-bounds + caps.",
        "Cover KV heuristics can mis-pair signature/stamp cells — review 表紙.",
        "Markdown is semantic reconstruction, not layout-lossless.",
    ]

    for ws in wb.worksheets:
        title = ws.title
        if wanted is not None and title.strip() not in wanted and title not in wanted:
            continue
        merges = len(list(ws.merged_cells.ranges))
        scan_r = max(max_rows + 40, 120)
        scan_c = max(max_cols + 20, 80)
        used_r, used_c = _used_bounds(ws, scan_rows=scan_r, scan_cols=scan_c)
        exp_r = min(used_r, max_rows)
        exp_c = min(used_c, max_cols)
        strategy = _sheet_strategy(title, merges, used_c)
        safe = _safe_name(title)

        html_path = out_dir / f"{safe}.html"
        md_path = out_dir / f"{safe}.md"
        html_path.write_text(sheet_to_html(ws, exp_r, exp_c, title), encoding="utf-8")

        md: str
        md_mode: str
        if strategy == "layout-asset":
            md = (
                f"# {title.strip()}\n\n"
                f"Layout sheet — see HTML preview `{html_path.name}`.\n\n"
                f"_strategy=layout-asset merges={merges}_\n"
            )
            md_mode = "layout-link"
            limitations.append(f"{title}: treated as layout-asset (HTML primary).")
        else:
            table_md = md_control_table(ws, title, max_rows=max_rows)
            if table_md:
                md = table_md
                md_mode = "control-table"
            elif strategy == "cover-kv":
                md = md_cover_kv(ws, exp_r, exp_c, title)
                md_mode = "cover-kv"
            else:
                md = md_blocks(ws, exp_r, exp_c, title)
                md_mode = "blocks"
        md_path.write_text(md, encoding="utf-8")

        results.append(
            {
                "sheet": title,
                "strategy": strategy,
                "md_mode": md_mode,
                "merges": merges,
                "used_approx": f"{used_r}x{used_c}",
                "exported": f"{exp_r}x{exp_c}",
                "html": str(html_path),
                "markdown": str(md_path),
            }
        )

    wb.close()
    report = {
        "format": "excel-doc-convert",
        "operation": "convert",
        "source": str(path),
        "output_dir": str(out_dir),
        "sheet_count": len(results),
        "sheets": results,
        "limitations": sorted(set(limitations)),
    }
    report_path = out_dir / "convert-report.json"
    report_path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    report["report"] = str(report_path)
    return report


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    sub = parser.add_subparsers(dest="cmd", required=True)

    p_class = sub.add_parser("classify", help="Summarize sheets, merges, strategies")
    p_class.add_argument("xlsx", type=Path)

    p_conv = sub.add_parser("convert", help="Write HTML + Markdown + convert-report.json")
    p_conv.add_argument("xlsx", type=Path)
    p_conv.add_argument("out_dir", type=Path)
    p_conv.add_argument(
        "--sheets",
        default="",
        help="Comma-separated sheet titles (default: all)",
    )
    p_conv.add_argument("--max-rows", type=int, default=120)
    p_conv.add_argument("--max-cols", type=int, default=40)

    args = parser.parse_args(argv)
    if args.cmd == "classify":
        if not args.xlsx.is_file():
            print(f"Error: file not found: {args.xlsx}", file=sys.stderr)
            return 2
        print(json.dumps(classify_workbook(args.xlsx), ensure_ascii=False, indent=2))
        return 0

    if args.cmd == "convert":
        if not args.xlsx.is_file():
            print(f"Error: file not found: {args.xlsx}", file=sys.stderr)
            return 2
        sheets = [s.strip() for s in args.sheets.split(",") if s.strip()] or None
        report = convert_workbook(
            args.xlsx,
            args.out_dir,
            sheet_filter=sheets,
            max_rows=args.max_rows,
            max_cols=args.max_cols,
        )
        print(json.dumps(report, ensure_ascii=False, indent=2))
        return 0

    return 2


if __name__ == "__main__":
    raise SystemExit(main())

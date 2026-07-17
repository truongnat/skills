#!/usr/bin/env python3
"""XLSX inspect / create / edit / validate with supported-lossless coverage."""

from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path
from typing import Any

_SKILLS_ROOT = Path(__file__).resolve().parents[2]
_COMMON = _SKILLS_ROOT / "office-common" / "python"
if not (_COMMON / "office_common").is_dir():
    raise SystemExit(f"office-common not found at {_COMMON}")
sys.path.insert(0, str(_COMMON))

from office_common.atomic import atomic_publish  # noqa: E402
from office_common.coverage import CoverageError, CoverageManifest  # noqa: E402
from office_common.fingerprint import semantic_fingerprint  # noqa: E402
from office_common.ooxml import XLSX_SUPPORTED_PREFIXES, inventory_ooxml  # noqa: E402
from office_common.tools import optional_tool_status  # noqa: E402


def _require_openpyxl():
    try:
        import openpyxl  # noqa: F401
    except ImportError as exc:
        raise SystemExit(
            "Missing openpyxl. Run this skill's scripts/setup.py, then invoke "
            "cli.py with the Python executable in this skill's .venv."
        ) from exc


def semantic_model(path: Path) -> dict[str, Any]:
    _require_openpyxl()
    from openpyxl import load_workbook

    wb = load_workbook(path, data_only=False)
    sheets = []
    for ws in wb.worksheets:
        cells = []
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None and (cell.data_type in (None, "n") or cell.data_type == "n"):
                    # skip empty cells without style-only noise in fingerprint
                    if cell.has_style and cell.value is None:
                        continue
                    continue
                cells.append(
                    {
                        "coord": cell.coordinate,
                        "value": cell.value,
                        "data_type": cell.data_type,
                        "number_format": cell.number_format,
                    }
                )
        sheets.append(
            {
                "title": ws.title,
                "dimensions": ws.dimensions,
                "cells": cells,
                "merged": [str(r) for r in ws.merged_cells.ranges],
            }
        )
    props = {
        "title": wb.properties.title,
        "creator": wb.properties.creator,
        "subject": wb.properties.subject,
    }
    return {"sheets": sheets, "properties": props, "sheetnames": list(wb.sheetnames)}


def inspect_xlsx(path: Path, operation: str = "inspect") -> CoverageManifest:
    manifest = CoverageManifest(
        format="xlsx",
        operation=operation,
        source=str(path),
        output=None,
    )
    status = "detected" if operation == "inspect" else "preserved"
    manifest.extend(inventory_ooxml(path, "xlsx", XLSX_SUPPORTED_PREFIXES, mark_status=status))

    try:
        model = semantic_model(path)
    except Exception as exc:  # noqa: BLE001
        if manifest.unsupported:
            manifest.notes.append(f"semantic_model_error={exc}")
            for tool in optional_tool_status():
                if tool.name == "libreoffice" and not tool.available:
                    manifest.add(
                        "tool:libreoffice",
                        "optional-check",
                        "skipped",
                        "formula recalculation not available",
                    )
            return manifest
        raise

    for sheet in model["sheets"]:
        manifest.add(f"sheet:{sheet['title']}", "worksheet", status, "worksheet title+cells")
        for cell in sheet["cells"]:
            manifest.add(
                f"sheet:{sheet['title']}/{cell['coord']}",
                "cell",
                status,
                f"value={cell['value']!r}",
            )
        for merged in sheet["merged"]:
            manifest.add(f"sheet:{sheet['title']}/merged:{merged}", "merged-range", status)

    for tool in optional_tool_status():
        if tool.name == "libreoffice" and not tool.available:
            manifest.add(
                "tool:libreoffice",
                "optional-check",
                "skipped",
                "formula recalculation not available",
            )
        elif tool.name == "libreoffice" and tool.available:
            manifest.add(
                "tool:libreoffice",
                "optional-check",
                "detected" if operation == "inspect" else "preserved",
                f"available at {tool.binary}",
            )

    manifest.notes.append(f"semantic_fingerprint={semantic_fingerprint(model)}")
    return manifest


def create_xlsx(spec: dict[str, Any], output: Path) -> CoverageManifest:
    _require_openpyxl()
    from openpyxl import Workbook

    manifest = CoverageManifest(
        format="xlsx",
        operation="create",
        source=None,
        output=str(output),
    )

    def writer(tmp: Path) -> None:
        wb = Workbook()
        # remove default sheet after building from spec
        default = wb.active
        sheets = spec.get("sheets") or [{"title": "Sheet1", "rows": []}]
        first = True
        for sheet_spec in sheets:
            title = sheet_spec.get("title") or "Sheet1"
            if first:
                ws = default
                ws.title = title
                first = False
            else:
                ws = wb.create_sheet(title)
            for r_idx, row in enumerate(sheet_spec.get("rows") or [], start=1):
                for c_idx, value in enumerate(row, start=1):
                    ws.cell(row=r_idx, column=c_idx, value=value)
                    manifest.add(
                        f"sheet:{title}/{ws.cell(row=r_idx, column=c_idx).coordinate}",
                        "cell",
                        "preserved",
                        f"value={value!r}",
                    )
            for merge in sheet_spec.get("merged") or []:
                ws.merge_cells(merge)
                manifest.add(f"sheet:{title}/merged:{merge}", "merged-range", "preserved")
            manifest.add(f"sheet:{title}", "worksheet", "preserved")
        props = spec.get("properties") or {}
        if props.get("title"):
            wb.properties.title = props["title"]
            manifest.add("docprops/title", "property", "preserved", props["title"])
        if props.get("creator"):
            wb.properties.creator = props["creator"]
            manifest.add("docprops/creator", "property", "preserved", props["creator"])
        wb.save(tmp)

    def validator(tmp: Path) -> None:
        post = inspect_xlsx(tmp, operation="validate")
        if post.unsupported:
            raise CoverageError("created file contains unsupported parts")
        # Mark package parts preserved in final manifest via re-inspect merge below
        model = semantic_model(tmp)
        expected_fp = semantic_fingerprint(
            {
                "sheets": [
                    {
                        "title": s.get("title") or "Sheet1",
                        "rows": s.get("rows") or [],
                        "merged": s.get("merged") or [],
                    }
                    for s in (spec.get("sheets") or [{"title": "Sheet1", "rows": []}])
                ]
            }
        )
        # Compare cell values loosely from written model
        written_values = {
            (sh["title"], c["coord"]): c["value"] for sh in model["sheets"] for c in sh["cells"]
        }
        for sheet_spec in spec.get("sheets") or [{"title": "Sheet1", "rows": []}]:
            title = sheet_spec.get("title") or "Sheet1"
            for r_idx, row in enumerate(sheet_spec.get("rows") or [], start=1):
                for c_idx, value in enumerate(row, start=1):
                    from openpyxl.utils import get_column_letter

                    coord = f"{get_column_letter(c_idx)}{r_idx}"
                    if written_values.get((title, coord)) != value:
                        raise CoverageError(
                            f"cell mismatch {title}!{coord}: wrote {written_values.get((title, coord))!r} expected {value!r}"
                        )
        _ = expected_fp  # reserved for future stricter structural fingerprint

    atomic_publish(output, writer, validator)
    final = inspect_xlsx(output, operation="create")
    # Re-tag data items as preserved for create success
    remapped = CoverageManifest(
        format="xlsx",
        operation="create",
        source=None,
        output=str(output),
        notes=final.notes,
    )
    for item in final.items:
        status = item.status
        if status == "detected":
            status = "preserved"
        remapped.add(item.path, item.kind, status, item.detail)
    remapped.require_full_coverage()
    return remapped


def edit_xlsx(path: Path, edits: dict[str, Any], output: Path) -> CoverageManifest:
    _require_openpyxl()
    from openpyxl import load_workbook

    pre = inspect_xlsx(path, operation="edit")
    if pre.unsupported:
        pre.require_full_coverage()

    def writer(tmp: Path) -> None:
        wb = load_workbook(path)
        for cell_edit in edits.get("set_cells") or []:
            sheet = cell_edit["sheet"]
            coord = cell_edit["coord"]
            wb[sheet][coord] = cell_edit["value"]
        wb.save(tmp)

    def validator(tmp: Path) -> None:
        post = inspect_xlsx(tmp, operation="validate")
        post.require_full_coverage()
        model = semantic_model(tmp)
        values = {
            (sh["title"], c["coord"]): c["value"] for sh in model["sheets"] for c in sh["cells"]
        }
        for cell_edit in edits.get("set_cells") or []:
            key = (cell_edit["sheet"], cell_edit["coord"])
            if values.get(key) != cell_edit["value"]:
                raise CoverageError(f"edit not applied for {key}")

    atomic_publish(output, writer, validator)
    final = inspect_xlsx(output, operation="edit")
    remapped = CoverageManifest(
        format="xlsx",
        operation="edit",
        source=str(path),
        output=str(output),
        notes=final.notes,
    )
    for item in final.items:
        status = "preserved" if item.status == "detected" else item.status
        remapped.add(item.path, item.kind, status, item.detail)
    for cell_edit in edits.get("set_cells") or []:
        remapped.add(
            f"sheet:{cell_edit['sheet']}/{cell_edit['coord']}",
            "cell",
            "transformed",
            f"set value={cell_edit['value']!r}",
        )
    remapped.require_full_coverage()
    return remapped


def validate_xlsx(path: Path) -> CoverageManifest:
    manifest = inspect_xlsx(path, operation="validate")
    remapped = CoverageManifest(
        format="xlsx",
        operation="validate",
        source=str(path),
        output=None,
        notes=manifest.notes,
    )
    for item in manifest.items:
        status = "preserved" if item.status == "detected" else item.status
        remapped.add(item.path, item.kind, status, item.detail)
    remapped.require_full_coverage()
    return remapped


def _load_json(path: Path) -> dict[str, Any]:
    return json.loads(path.read_text(encoding="utf-8"))


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="XLSX supported-lossless office skill CLI")
    sub = parser.add_subparsers(dest="cmd", required=True)

    p_insp = sub.add_parser("inspect", help="Inventory package parts and cell data")
    p_insp.add_argument("input", type=Path)
    p_insp.add_argument("--manifest", type=Path, help="Write coverage JSON")

    p_create = sub.add_parser("create", help="Create xlsx from JSON spec")
    p_create.add_argument("spec", type=Path)
    p_create.add_argument("output", type=Path)
    p_create.add_argument("--manifest", type=Path)

    p_edit = sub.add_parser("edit", help="Edit cells from JSON edits")
    p_edit.add_argument("input", type=Path)
    p_edit.add_argument("edits", type=Path)
    p_edit.add_argument("output", type=Path)
    p_edit.add_argument("--manifest", type=Path)

    p_val = sub.add_parser("validate", help="Validate supported-lossless coverage")
    p_val.add_argument("input", type=Path)
    p_val.add_argument("--manifest", type=Path)

    args = parser.parse_args(argv)
    try:
        if args.cmd == "inspect":
            manifest = inspect_xlsx(args.input)
        elif args.cmd == "create":
            manifest = create_xlsx(_load_json(args.spec), args.output)
        elif args.cmd == "edit":
            manifest = edit_xlsx(args.input, _load_json(args.edits), args.output)
        elif args.cmd == "validate":
            manifest = validate_xlsx(args.input)
        else:
            raise SystemExit(f"unknown command {args.cmd}")
        if args.manifest:
            manifest.write_json(args.manifest)
        print(json.dumps(manifest.to_dict(), indent=2, sort_keys=True))
        return 0
    except CoverageError as exc:
        print(f"COVERAGE_ERROR: {exc}", file=sys.stderr)
        return 2
    except Exception as exc:  # noqa: BLE001
        print(f"ERROR: {exc}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())

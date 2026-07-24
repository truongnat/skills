from __future__ import annotations

import json
import subprocess
import sys
from pathlib import Path

from openpyxl import Workbook

ROOT = Path(__file__).resolve().parents[1]
CLI = ROOT / "skills" / "excel-doc-convert" / "scripts" / "cli.py"
PYTHON = sys.executable


def _run(args: list[str]) -> subprocess.CompletedProcess[str]:
    return subprocess.run(
        [PYTHON, str(CLI), *args],
        cwd=str(ROOT),
        text=True,
        capture_output=True,
        check=False,
    )


def _make_merged_doc(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "表紙"
    ws["C2"] = "画面設計書"
    ws.merge_cells("C2:H4")
    ws["C6"] = "機能ID"
    ws["G6"] = "FBD13001"
    ws.merge_cells("G6:J6")
    ws["C7"] = "機能名"
    ws["G7"] = "製造指図書"

    detail = wb.create_sheet("画面項目詳細")
    detail["A1"] = "№"
    detail["C1"] = "コントロール名"
    detail.merge_cells("C1:E1")
    detail["F1"] = "コントロールID"
    detail.merge_cells("F1:H1")
    detail["A2"] = "1"
    detail["C2"] = "検索(F10)"
    detail.merge_cells("C2:E2")
    detail["F2"] = "F10"
    detail.merge_cells("F2:H2")
    detail["A3"] = "2"
    detail["C3"] = "拠点"
    detail.merge_cells("C3:E3")
    detail["F3"] = "lblBase"
    detail.merge_cells("F3:H3")

    layout = wb.create_sheet("画面レイアウト")
    layout["A1"] = "wire"
    layout.merge_cells("A1:D8")

    wb.save(path)


def test_classify_and_convert_merged_doc(tmp_path: Path) -> None:
    xlsx = tmp_path / "sample.xlsx"
    out = tmp_path / "out"
    _make_merged_doc(xlsx)

    classified = _run(["classify", str(xlsx)])
    assert classified.returncode == 0, classified.stderr
    data = json.loads(classified.stdout)
    assert data["format"] == "excel-doc-convert"
    assert data["operation"] == "classify"
    assert data["merge_total"] >= 3
    titles = {s["title"] for s in data["sheets"]}
    assert "表紙" in titles
    assert "画面レイアウト" in titles

    converted = _run(
        ["convert", str(xlsx), str(out), "--sheets", "表紙,画面項目詳細,画面レイアウト"]
    )
    assert converted.returncode == 0, converted.stderr
    report = json.loads(converted.stdout)
    assert report["operation"] == "convert"
    assert (out / "convert-report.json").is_file()
    assert (out / "表紙.html").is_file()
    assert (out / "表紙.md").is_file()
    cover = (out / "表紙.md").read_text(encoding="utf-8")
    assert "FBD13001" in cover
    detail_md = (out / "画面項目詳細.md").read_text(encoding="utf-8")
    assert "lblBase" in detail_md
    assert "F10" in detail_md
    layout_md = (out / "画面レイアウト.md").read_text(encoding="utf-8")
    assert "layout-asset" in layout_md or "HTML" in layout_md
    assert any("Shapes" in x or "layout" in x.lower() for x in report["limitations"])
